# -*- coding: utf-8 -*-
"""
招行 U-BANK - 日记账流程：复用公共登录能力并进入日记账页面
支持两种模式：
  1. 查询模式（默认）：登录U-BANK -> 提取数据 -> 保存JSON -> 写入Excel -> 退出
  2. 写入模式：从已有JSON文件读取数据 -> 追加写入Excel（不启动U-BANK）
用法：
  python 日记账.py              # 查询模式（完整流程）
  python 日记账.py --write      # 写入模式（仅JSON->Excel）
  python 日记账.py -w json文件   # 写入模式（指定JSON文件）
"""
import time
import os
import sys
import re
import json
import argparse
from datetime import datetime
from pywinauto import Desktop
import win32gui

# 将项目根目录加入模块搜索路径，复用公共模块 ubank_common
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from ubank_common import (
    open_ubank, login_ubank, wait_for_main_window,
    click_control_by_names, close_with_confirm, click_at,
    ensure_dir, capture_window_screenshot, SCREENSHOT_DIR,
)
from pywinauto.keyboard import send_keys


def _ts():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _screenshot(main_win, tag):
    ensure_dir(SCREENSHOT_DIR)
    path = os.path.join(SCREENSHOT_DIR, f"rijizhang_{tag}_{_ts()}.png")
    capture_window_screenshot(main_win, path)
    return path


def enter_transaction_flow(main_win):
    """进入'交易明细':通过控件名称查找并点击"""
    time.sleep(6)  # 首页渲染等待
    # 尝试多种候选控件名，覆盖不同版本/语言环境
    candidates = ["交易明细", "交易流水", "日记账", "账户明细"]
    if click_control_by_names(main_win, candidates):
        print("已通过控件名称点击进入交易明细页面")
    else:
        print("控件名称未命中，回退到坐标点击")
        rect = main_win.rectangle()
        # 点击右侧自定义工作台区域的"交易明细"标签页
        x = int(rect.left + rect.width() * 0.855)
        y = int(rect.top + rect.height() * 0.285)
        click_at(x, y)
        print(f"已使用坐标点击: ({x}, {y})")
    time.sleep(2)
    return True


def set_date_range_and_query(main_win):
    """按坐标直接选择本月1日到今天，点击查询"""
    time.sleep(6)  # 交易查询页渲染等待

    today = datetime.now()
    start_date = today.replace(day=1).strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")

    rect = main_win.rectangle()
    w, h = rect.width(), rect.height()

    def click_field_and_type(px, py, value):
        x = int(rect.left + w * px)
        y = int(rect.top + h * py)
        click_at(x, y)
        time.sleep(0.3)
        send_keys("^a")
        time.sleep(0.1)
        send_keys("{DEL}")
        time.sleep(0.1)
        send_keys(value)
        time.sleep(0.2)
        send_keys("{ENTER}")
        time.sleep(0.3)
        print(f"写入日期: {value}")

    click_field_and_type(0.202, 0.318, start_date)
    click_field_and_type(0.295, 0.318, end_date)

    _screenshot(main_win, "after_date")

    # 查询按钮
    time.sleep(1.5)  # 日期回填后稍等
    qx = int(rect.left + w * 0.952)
    qy = int(rect.top + h * 0.274)
    click_at(qx, qy)
    print("已点击查询")
    time.sleep(3)  # 等查询结果返回


def _collect_control_texts(control, depth=0):
    """递归收集控件树中所有有文本内容的叶子控件"""
    results = []
    if depth > 15:  # 防止无限递归
        return results
    try:
        name = ""
        ctrl_type = ""
        rect_str = ""

        try:
            name = control.element_info.name if hasattr(control.element_info, "name") else ""
            name = str(name).strip() if name else ""
        except Exception:
            pass

        try:
            ctrl_type = control.element_info.control_type if hasattr(control.element_info, "control_type") else ""
        except Exception:
            pass

        try:
            r = control.rectangle()
            rect_str = f"{r.left},{r.top},{r.right},{r.bottom}"
        except Exception:
            pass

        # 记录有效文本（排除空字符串和过长的无意义内容）
        if name and len(name) > 0 and len(name) < 200 and name not in ("", " ", "\n"):
            results.append({
                "name": name,
                "type": str(ctrl_type),
                "rect": rect_str,
                "depth": depth,
            })

        # 递归子控件
        try:
            for child in control.children():
                results.extend(_collect_control_texts(child, depth + 1))
        except Exception:
            pass
    except Exception:
        pass
    return results


def _parse_amount(val_str):
    """解析金额字符串为浮点数"""
    if not val_str or val_str.strip() in ("", "-", "--"):
        return None
    try:
        s = val_str.strip().replace(",", "").replace(" ", "")
        # 处理带+号的金额
        s = s.replace("+", "")
        return float(s)
    except (ValueError, TypeError):
        return None


def _is_valid_date(text):
    """判断文本是否为有效的交易日期格式"""
    if not text:
        return False
    # 匹配 2026-04-22 或 2026.4.22 等格式
    pattern = r'^\d{4}[-./]\d{1,2}[-./]\d{1,2}'
    return bool(re.match(pattern, text.strip()))


def _is_menu_or_ui_noise(text):
    """判断文本是否为菜单/导航/工具栏等干扰项"""
    noise_patterns = [
        # 左侧菜单项
        "首页", "工作台", "账户管理", "付款管理", "薪酬管理", "收款管理",
        "现金管理", "跨境金融", "财富管理", "票据管理", "回单管理",
        "管理驾驶舱", "报表中心", "预警中心", "集团架构管理",
        # 右侧工具栏
        "导出", "缩放", "帮助", "演示", "切换", "旧版", "App",
        # 底部分页/统计
        "选择全部", "共.*笔", "人民币.*出账.*入账",
        # 英文菜单ID
        "collection-center", "cash-management", "cross-financial",
        "new-wealth-management", "bill-management-center",
        "receipt-management-center", "management-cockpit",
        "new-report-center", "early-warning-center",
        "architecture-management",
        # 其他UI噪声
        "planb-down", "left", "right",
    ]
    t = text.strip()
    for p in noise_patterns:
        if re.search(p, t, re.IGNORECASE):
            return True
    return False


def extract_transaction_data(main_win):
    """
    提取交易明细表格数据，按Excel模板格式输出：
      日期 | 摘要 | 收（付）款人 | 收入 | 支出 | 余额 | 内部明细
    同时保存JSON和Excel文件
    """
    print("开始提取交易明细表格数据...")
    time.sleep(2)

    all_texts = _collect_control_texts(main_win)
    print(f"共收集到 {len(all_texts)} 个文本控件")

    # 表格列名关键字（U-BANK界面上的表头）
    header_keywords = ["交易日期", "交易金额", "余额", "收(付)方名称",
                       "交易类型", "摘要", "标签"]

    # 第一步：定位表头，找到包含列名的区域
    headers_found = []
    for item in all_texts:
        for kw in header_keywords:
            if kw in item["name"]:
                headers_found.append(item)
                break

    if len(headers_found) < 2:
        print("未能识别到足够多的表头列，尝试按坐标聚类提取...")
        result = _extract_by_coordinate_clustering(all_texts)
        _save_to_excel(result)
        return result

    print(f"识别到表头列数: {len(headers_found)}")

    # 第二步：用表头的Y坐标范围确定表格行的Y区间
    header_y_values = []
    for h in headers_found:
        parts = h["rect"].split(",")
        if len(parts) == 4:
            try:
                header_y_values.append(int(parts[1]))  # top
            except (ValueError, IndexError):
                pass

    if not header_y_values:
        print("无法获取表头Y坐标")
        result = {"提取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "总记录数": 0, "交易明细": []}
        _save_to_excel(result)
        return result

    max_header_y = max(header_y_values)

    # 第三步：筛选出表头下方、可能是数据的文本项
    data_candidates = []
    for item in all_texts:
        parts = item["rect"].split(",")
        if len(parts) != 4 or item["depth"] > 12:
            continue
        try:
            y_top = int(parts[1])
            if y_top > max_header_y + 10:
                data_candidates.append(item)
        except (ValueError, IndexError):
            continue

    print(f"表头下方候选数据项: {len(data_candidates)} 个")

    # 第四步：按Y坐标分组为行
    rows_dict = {}
    for item in data_candidates:
        parts = item["rect"].split(",")
        if len(parts) != 4:
            continue
        try:
            y_top = int(parts[1])
            row_key = round(y_top / 8) * 8  # 8px容差
            if row_key not in rows_dict:
                rows_dict[row_key] = []
            rows_dict[row_key].append(item)
        except (ValueError, IndexError):
            continue

    sorted_row_keys = sorted(rows_dict.keys())

    # 第五步：每行内按X排序，映射到Excel模板列名
    # U-BANK表格列顺序（从截图看）：交易日期, 交易金额, 余额, 收(付)方名称, 收(付)方账号, 交易类型, 标签
    # 映射到Excel模板：         日期,     [金额],   余额, 收（付）款人,  [账号],      摘要/类型,  内部明细
    records = []
    for row_key in sorted_row_keys:
        row_items = rows_dict[row_key]
        row_items.sort(key=lambda x: int(x["rect"].split(",")[0]) if len(x["rect"].split(",")) == 4 else 0)
        texts_in_row = [item["name"] for item in row_items]

        # 过滤操作按钮文字和UI噪声
        clean_texts = []
        for t in texts_in_row:
            t_stripped = t.strip()
            if any(kw in t for kw in ["回单打印", "转发至微信"]):
                continue
            if _is_menu_or_ui_noise(t_stripped):
                continue
            clean_texts.append(t_stripped)

        if not clean_texts:
            continue

        # 检查第一列是否为有效日期格式（过滤掉菜单/工具栏等干扰行）
        if not _is_valid_date(clean_texts[0]):
            print(f"  跳过非数据行(首列非日期): {clean_texts[0][:30]}")
            continue

        # 构建一行记录，按位置匹配
        rec = {
            "日期": "",
            "摘要": "",
            "收（付）款人": "",
            "收入": "",
            "支出": "",
            "余额": "",
            "内部明细": "",
        }

        idx = 0
        for text in clean_texts:
            if idx >= 7:
                break
            if idx == 0:
                rec["日期"] = text
            elif idx == 1:
                amount_val = _parse_amount(text)
                if amount_val is not None:
                    if amount_val >= 0:
                        rec["收入"] = text.replace("+", "") if text.startswith("+") else text
                    else:
                        rec["支出"] = text
                else:
                    rec["支出"] = text  # 无法解析时原样放入
            elif idx == 2:
                rec["余额"] = text
            elif idx == 3:
                rec["收（付）款人"] = text
            elif idx == 5:
                rec["摘要"] = text  # 交易类型作为摘要
            elif idx == 6:
                rec["内部明细"] = text  # 标签作为内部明细
            idx += 1

        records.append(rec)

    # 第六步：构建结果
    result = {
        "提取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "总记录数": len(records),
        "交易明细": records,
    }

    # 保存JSON
    ensure_dir(SCREENSHOT_DIR)
    json_path = os.path.join(SCREENSHOT_DIR, f"rijizhang_data_{_ts()}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"JSON已保存: {json_path}")

    # 追加写入现有Excel文件（去重，只写新增记录）
    append_to_existing_excel(result)

    print(f"交易明细数据已提取: 共 {len(records)} 条记录")
    if records:
        print("--- 交易记录摘要 ---")
        for i, rec in enumerate(records[:8], 1):
            date = rec.get("日期", "-")
            income = rec.get("收入", "-")
            expense = rec.get("支出", "-")
            balance = rec.get("余额", "-")
            name = rec.get("收（付）款人", "-")[:15]
            desc = rec.get("摘要", "-")[:10]
            print(f"  [{i}] {date} | 收:{income} 支:{expense} | 余:{balance} | {name} | {desc}")
        if len(records) > 8:
            print(f"  ... 还有 {len(records)-8} 条记录")

    return result


def _norm_amount(val):
    """标准化金额字符串，用于去重对比（去除逗号、正负号差异）"""
    if val is None or str(val).strip() in ("", "-", "--", "NaN", "None"):
        return ""
    s = str(val).strip().replace(",", "").replace(" ", "")
    s = s.lstrip("-").lstrip("+")
    return s


def _clean_date(raw_date):
    """清洗日期：只保留年-月-日部分，去掉时分秒"""
    import re
    # 匹配 2026-04-04 或 2026/4/4 等格式的日期部分
    m = re.match(r'(\d{4}[-./]\d{1,2}[-./]\d{1,2})', raw_date.strip())
    return m.group(1) if m else raw_date.strip()


def _to_positive(raw_exp):
    """支出金额转正数（去掉负号）"""
    if not raw_exp or raw_exp.strip() in ("", "-", "--", "None"):
        return ""
    s = raw_exp.strip().lstrip("-")
    return s


def append_to_existing_excel(result):
    """将提取结果追加写入现有的银行日记账Excel文件（只操作A~G列，H/I列不动）"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    records = result.get("交易明细", [])
    if not records:
        print("无数据需要追加")
        return

    # Excel文件路径
    excel_path = os.path.join(PROJECT_ROOT, "4--北京示例付款单位B-银行日记账.xlsx")
    if not os.path.exists(excel_path):
        print(f"目标Excel不存在: {excel_path}")
        _save_to_excel(result)
        return

    wb = load_workbook(excel_path)
    ws = wb.active
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 收集已有记录的key，用于去重：(日期前10位 + 收入 + 支出 + 余额)
    # 不依赖摘要字段，因为Excel中摘要是人工标签，JSON中是U-BANK原始值
    existing_keys = set()
    last_data_row = 3  # 表头在第3行，数据从第4行开始
    for r in range(4, ws.max_row + 1):
        date_val = ws.cell(row=r, column=1).value
        if date_val is not None and str(date_val).strip():
            last_data_row = r
            d = str(date_val).strip()[:10]
            inc = _norm_amount(ws.cell(row=r, column=4).value)
            exp = _norm_amount(ws.cell(row=r, column=5).value)
            bal = _norm_amount(ws.cell(row=r, column=6).value)
            existing_keys.add((d, inc, exp, bal))

    print(f"现有数据最后行: {last_data_row}, 已有记录数: {len(existing_keys)}")

    # 筛选出真正新增的记录并追加
    append_count = 0
    write_row = last_data_row + 1
    for rec in records:
        rec_date = str(rec.get("日期", "") or "").strip()[:10]
        rec_inc = _norm_amount(rec.get("收入", ""))
        rec_exp = _norm_amount(rec.get("支出", ""))
        rec_bal = _norm_amount(rec.get("余额", ""))
        key = (rec_date, rec_inc, rec_exp, rec_bal)

        if key in existing_keys:
            print(f"  跳过已存在: {rec_date} | 支出:{rec_exp}")
            continue

        # 清洗数据后再写入，使格式与人工填写一致
        raw_date = str(rec.get("日期", "") or "").strip()
        clean_date = _clean_date(raw_date)
        raw_exp = str(rec.get("支出", "") or "").strip()
        clean_exp = _to_positive(raw_exp)

        values = [
            clean_date,
            rec.get('摘要', ''),
            rec.get('收（付）款人', ''),
            rec.get('收入', '') if rec.get('收入') else '',
            clean_exp,
            rec.get('余额', ''),
            rec.get('内部明细', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=write_row, column=col_idx, value=val)
            cell.border = thin_border
        write_row += 1
        append_count += 1

    wb.save(excel_path)
    print(f"Excel追加完成: 新增 {append_count} 条记录 -> {excel_path}")


def _save_to_excel(result):
    """将提取结果按Excel模板格式保存为xlsx文件"""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    records = result.get("交易明细", [])
    ensure_dir(SCREENSHOT_DIR)
    xlsx_path = os.path.join(SCREENSHOT_DIR, f"rijizhang_data_{_ts()}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "银行日记账"

    # 表头样式
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 写标题行
    ws['A1'] = '银行日记账'
    ws.merge_cells('A1:G1')
    ws['A1'].font = Font(bold=True, size=14)

    # 写信息行
    ws['A2'] = '公司：北京示例付款单位B科技有限公司'
    ws.merge_cells('A2:C2')
    ws['D2'] = '开户行：招商银行北京经济技术开发区科技金融支行'
    ws.merge_cells('D2:F2')
    ws['G2'] = '账号；000000000000000000'

    # 写表头（第3行）
    excel_headers = ['日期', '摘要', '收（付）款人', '收入', '支出', '余额', '内部明细']
    for col_idx, h in enumerate(excel_headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 写数据行（从第4行开始）
    for row_idx, rec in enumerate(records, 4):
        values = [
            rec.get('日期', ''),
            rec.get('摘要', ''),
            rec.get('收（付）款人', ''),
            rec.get('收入', '') if rec.get('收入') else '',
            rec.get('支出', '') if rec.get('支出') else '',
            rec.get('余额', ''),
            rec.get('内部明细', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # 设置列宽
    col_widths = {'A': 16, 'B': 30, 'C': 28, 'D': 14, 'E': 14, 'F': 14, 'G': 12}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(xlsx_path)
    print(f"Excel已保存: {xlsx_path}")


def _extract_by_coordinate_clustering(all_texts):
    """备用方案：当无法识别表头时，通过坐标聚类提取数据"""
    print("使用坐标聚类方案提取数据...")

    # 过滤掉明显不是表格数据的文本
    skip_keywords = [
        "首页", "网银设置", "企业APP", "自助申请", "财务管理", "转账支付",
        "代发代扣", "现金管理", "投资理财", "特色存款", "更多功能",
        "单账号查看", "多账号查看", "查询", "重置", "收起",
        "导出全部", "打印全部", "分享全部", "选择全部", "共", "笔",
        "人民币", "出账", "入账", "条页", "U-BANK用户名",
        "上次登录时间", "网银系统时间", "当前功能", "账务管理",
        "回单管理中心", "支持近五年历史交易查询", "预设凭证查询模板",
    ]

    filtered = []
    for item in all_texts:
        name = item.get("name", "")
        # 跳过无关文本
        if any(kw in name for kw in skip_keywords):
            continue
        # 跳过纯数字或极短无意义文本
        if len(name.strip()) <= 1:
            continue
        # 跳过过长文本
        if len(name) > 100:
            continue
        filtered.append(item)

    print(f"过滤后剩余候选: {len(filtered)} 个")

    # 按Y坐标分行
    rows_dict = {}
    for item in filtered:
        parts = item.get("rect", "").split(",")
        if len(parts) != 4:
            continue
        try:
            y_top = int(parts[1])
            row_key = round(y_top / 8) * 8
            if row_key not in rows_dict:
                rows_dict[row_key] = []
            rows_dict[row_key].append(item)
        except (ValueError, IndexError):
            continue

    sorted_rows = sorted(rows_dict.keys())
    records = []
    for row_key in sorted_rows:
        items = rows_dict[row_key]
        items.sort(key=lambda x: int(x["rect"].split(",")[0]) if len(x["rect"].split(",")) == 4 else 0)
        row_texts = [item["name"] for item in items]
        # 过滤掉操作按钮文字
        clean_texts = [t for t in row_texts if not any(kw in t for kw in ["回单打印", "转发至微信"])]
        if clean_texts and len(clean_texts) >= 2:  # 至少两列才视为有效行
            records.append(clean_texts)

    result = {
        "提取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "总记录数": len(records),
        "说明": "坐标聚类模式-未识别到标准表头",
        "交易明细": [{"行数据": row} for row in records],
    }

    ensure_dir(SCREENSHOT_DIR)
    json_path = os.path.join(SCREENSHOT_DIR, f"rijizhang_data_{_ts()}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"聚类提取完成: {len(records)} 条记录 -> {json_path}")
    return result


def _get_hwnd(win):
    try:
        return int(win.handle)
    except Exception:
        try:
            return int(win.element_info.handle)
        except Exception:
            return 0


def is_hwnd_alive(hwnd):
    if not hwnd:
        return False
    try:
        return bool(win32gui.IsWindow(hwnd))
    except Exception:
        return False


def finalize_exit(main_hwnd, timeout=8):
    """轮询等待主窗 HWND 消失；仍在则复用 close_with_confirm 再走一次"""
    end_ts = time.time() + timeout
    while time.time() < end_ts:
        if not is_hwnd_alive(main_hwnd):
            print("检测到 U-BANK 主窗已关闭")
            return True
        time.sleep(0.5)
    print("主窗仍存活，复用 close_with_confirm 再执行一次退出")
    try:
        desktop = Desktop(backend="uia")
        alive = desktop.window(handle=main_hwnd)
        if alive.exists():
            close_with_confirm(alive)
    except Exception as e:
        print(f"二次退出失败: {e}")
    time.sleep(2)
    return not is_hwnd_alive(main_hwnd)


def screenshot_after_exit():
    try:
        from PIL import ImageGrab
    except Exception as e:
        print(f"退出后截图失败: {e}")
        return None
    ensure_dir(SCREENSHOT_DIR)
    path = os.path.join(SCREENSHOT_DIR, f"rijizhang_after_exit_{_ts()}.png")
    hwnd = win32gui.GetForegroundWindow()
    if not hwnd:
        return None
    l, t, r, b = win32gui.GetWindowRect(hwnd)
    if r - l < 50 or b - t < 50:
        return None
    ImageGrab.grab(bbox=(l, t, r, b)).save(path)
    print(f"退出后截图已保存: {path}")
    return path


def main():
    open_ubank()
    if not login_ubank():
        print("登录失败，终止执行")
        return

    desktop = Desktop(backend="uia")
    print("等待主界面加载...")
    main_win = wait_for_main_window(desktop)
    if not main_win:
        print("未找到主界面，终止执行")
        return

    main_hwnd = _get_hwnd(main_win)
    print(f"主窗 HWND: {main_hwnd}")

    if enter_transaction_flow(main_win):
        print("已进入交易明细页面")
        set_date_range_and_query(main_win)
        extract_transaction_data(main_win)
    else:
        print("未成功进入交易明细页面")

    _screenshot(main_win, "before_exit")

    close_with_confirm(main_win)
    finalize_exit(main_hwnd)

    screenshot_after_exit()


def load_json_and_write(json_path=None):
    """从JSON文件读取数据并追加写入Excel（不启动U-BANK）"""
    if json_path is None:
        ensure_dir(SCREENSHOT_DIR)
        json_files = sorted(
            [f for f in os.listdir(SCREENSHOT_DIR) if f.startswith("rijizhang_data_") and f.endswith(".json")],
            reverse=True,
        )
        if not json_files:
            print("未找到任何JSON数据文件")
            return False
        json_path = os.path.join(SCREENSHOT_DIR, json_files[0])
        print(f"使用最新JSON文件: {json_files[0]}")

    if not os.path.exists(json_path):
        print(f"JSON文件不存在: {json_path}")
        return False

    with open(json_path, "r", encoding="utf-8") as f:
        result = json.load(f)

    records = result.get("交易明细", [])
    print(f"从JSON加载了 {len(records)} 条记录")
    append_to_existing_excel(result)


def main():
    parser = argparse.ArgumentParser(description="招行U-BANK日记账提取工具")
    parser.add_argument("-w", "--write", nargs="?", const="", default=None, metavar="JSON_PATH",
                        help="写入模式：从JSON读取数据写入Excel，可指定JSON路径")
    args = parser.parse_args()

    # 写入模式：不启动U-BANK，直接JSON -> Excel
    if args.write is not None:
        json_file = args.write if args.write else None
        load_json_and_write(json_file)
        return

    # 查询模式（默认）：完整U-BANK自动化流程
